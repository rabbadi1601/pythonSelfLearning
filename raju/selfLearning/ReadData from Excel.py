import openpyxl

def get_data_from_excel(path,sheet_name):

    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2,total_rows+1):
        row_list = []
        for c in range(1,total_cols+1):
            row_list.append(sheet.cell(row=r,column=c).value)
        final_list.append(row_list)

    return final_list

excelData=get_data_from_excel("../configurations/TutorialsNinja.xlsx","LoginTest")

#excelData=get_data_from_excel("../configurations/sales_data.xlsx","2020")

for value in excelData:
    print(" email :{} and password :{}".format(value[0],value[1]))
